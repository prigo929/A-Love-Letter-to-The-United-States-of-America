#!/usr/bin/env python3
"""Build lib/data/ports.json by extracting 2022 tonnage data from the USACE excel
sheet and adding geographic coordinates for the major deepwater ports of the US.
"""
import os
import json
import openpyxl

SRC = "ASSETS/GeoJSON/Ports army corps of engineers data 2022.xlsx"
OUT = "lib/data/ports.json"

PORT_COORDS = {
    "Houston Port Authority, TX": [-95.10, 29.75],
    "South Louisiana, LA, Port of": [-90.50, 30.00],
    "Corpus Christi, TX": [-97.40, 27.84],
    "New York, NY & NJ": [-74.02, 40.67],
    "Port of Long Beach, CA": [-118.21, 33.74],
    "New Orleans, LA": [-90.08, 29.93],
    "Beaumont, TX": [-94.08, 30.08],
    "Port of Greater Baton Rouge, LA": [-91.19, 30.43],
    "Virginia, VA, Port of": [-76.30, 36.88],
    "Lake Charles Harbor District, LA": [-93.22, 30.22],
    "Port of Los Angeles, CA": [-118.26, 33.73],
    "Plaquemines Port District, LA": [-89.96, 29.50],
    "Port of Savannah, GA": [-81.13, 32.12],
    "Mobile, AL": [-88.04, 30.70],
    "Port Arthur, TX": [-93.93, 29.86],
    "Baltimore, MD": [-76.58, 39.27],
    "Texas City, TX": [-94.90, 29.38],
    "Philadelphia Regional Port, PA": [-75.14, 39.89],
    "Port Freeport, TX": [-95.35, 28.94],
    "Duluth-Superior, MN and WI": [-92.09, 46.78],
    "Tampa Port Authority, FL": [-82.45, 27.90],
    "Port of Charleston, SC": [-79.91, 32.79],
    "Port Everglades, FL": [-80.12, 26.08],
    "Port of Pascagoula, MS": [-88.55, 30.36],
    "Richmond, CA": [-122.37, 37.92],
    "Port of Portland, OR": [-122.75, 45.64],
    "Tacoma, WA": [-122.41, 47.27],
    "Seattle, WA": [-122.34, 47.60],
    "Port of Oakland, CA": [-122.31, 37.80],
    "Jacksonville, FL": [-81.56, 30.38],
    "Pittsburgh, PA Port of": [-80.00, 40.44],
    "Port of Kalama, WA": [-122.84, 46.01],
    "Galveston, TX": [-94.79, 29.31],
    "Boston, MA": [-71.03, 42.35],
}

def main():
    print("Opening USACE Excel sheet...", flush=True)
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['Port_by_Tons']
    
    ports_list = []
    # Data rows start at row 6
    row_idx = 6
    while True:
        rank = ws.cell(row=row_idx, column=1).value
        if rank is None:
            break
        name = ws.cell(row=row_idx, column=2).value
        total = ws.cell(row=row_idx, column=3).value
        domestic = ws.cell(row=row_idx, column=4).value
        foreign = ws.cell(row=row_idx, column=5).value
        imports = ws.cell(row=row_idx, column=6).value
        exports = ws.cell(row=row_idx, column=7).value
        
        if name in PORT_COORDS:
            lng, lat = PORT_COORDS[name]
            
            # Clean up the name for user-facing displays
            clean_name = name.replace("Port of ", "").replace("Port Authority", "").replace("Harbor District", "").replace("Regional Port", "").replace("Port District", "").replace("Port Corp", "").strip()
            
            ports_list.append({
                "id": str(rank),
                "rank": int(rank),
                "name": clean_name,
                "fullName": name,
                "lng": lng,
                "lat": lat,
                "total": int(total) if total is not None else 0,
                "domestic": int(domestic) if domestic is not None else 0,
                "foreign": int(foreign) if foreign is not None else 0,
                "imports": int(imports) if imports is not None else 0,
                "exports": int(exports) if exports is not None else 0,
            })
        row_idx += 1
        
    out = {"ports": ports_list}
    print(f"Matched {len(ports_list)} ports out of {row_idx - 6} spreadsheet entries.", flush=True)
    
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Wrote {OUT} — {os.path.getsize(OUT)/1024:.1f} KB", flush=True)

if __name__ == "__main__":
    main()
